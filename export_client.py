"""
Client d'export pour écrire les données des tickets Zendesk en CSV ou Excel.

Ce module gère l'export des données normalisées vers des fichiers locaux.
"""

import csv
import os
from typing import List, Dict, Optional
import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# Fichier unique enrichi (base pour Looker Studio), mis à jour par import complet + incrémental
ALL_TICKETS_FILENAME = "tickets_all.csv"

# En-têtes par défaut pour les fichiers (alignés sur tickets_all.csv)
DEFAULT_HEADERS = [
    "Ticket ID",
    "Ticket status",
    "Ticket group",
    "Assignee name",
    "Ticket created - Date",
    "Ticket solved - Date",
    "Ticket brand",
    "Ticket channel",
    "Ticket form",
    "Ticket priority",
    "Ticket subject",
    "Ticket type",
    "Requester name",
    "Ticket organisation name",
    "Tickets",
]


class ExportClient:
    """Client pour exporter les tickets vers CSV ou Excel."""
    
    def __init__(self, output_dir: str = "exports", file_format: str = "csv"):
        """
        Initialise le client d'export.
        
        Args:
            output_dir: Répertoire où sauvegarder les fichiers
            file_format: Format de fichier ('csv' ou 'xlsx')
        """
        self.output_dir = Path(output_dir)
        self.file_format = file_format.lower()
        
        # Créer le répertoire s'il n'existe pas
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        if self.file_format not in ['csv', 'xlsx']:
            raise ValueError(f"Format non supporté: {file_format}. Utilisez 'csv' ou 'xlsx'")

    def get_all_tickets_path(self) -> Path:
        """Chemin du fichier cumulatif tickets_all.csv (base enrichie pour Looker Studio)."""
        return self.output_dir / ALL_TICKETS_FILENAME

    def _get_filename(self, prefix: str = "tickets") -> str:
        """
        Génère un nom de fichier avec timestamp.
        
        Args:
            prefix: Préfixe du nom de fichier
            
        Returns:
            str: Nom de fichier complet avec chemin
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = "csv" if self.file_format == "csv" else "xlsx"
        filename = f"{prefix}_{timestamp}.{extension}"
        return str(self.output_dir / filename)
    
    def _ticket_to_row(self, ticket: Dict) -> List:
        """
        Convertit un ticket en liste de valeurs pour une ligne (ordre = DEFAULT_HEADERS).
        """
        return [
            str(ticket.get("ticket_id", "")),
            ticket.get("status", ""),
            ticket.get("ticket_group", ""),
            ticket.get("assignee_name", ""),
            ticket.get("created_at", ""),
            ticket.get("updated_at", ""),
            ticket.get("brand", ""),
            ticket.get("via", ""),
            ticket.get("ticket_form", ""),
            ticket.get("priority", ""),
            ticket.get("subject", ""),
            ticket.get("type", ""),
            ticket.get("requester_name", ""),
            ticket.get("organisation_name", ""),
            ticket.get("tickets", "1"),
        ]
    
    def export_to_csv(self, tickets: List[Dict], filename: Optional[str] = None, append: bool = False) -> str:
        """
        Exporte les tickets vers un fichier CSV.
        
        Args:
            tickets: Liste de tickets normalisés
            filename: Nom de fichier personnalisé (optionnel)
            append: Si True, ajoute les données au fichier existant (sans en-têtes)
            
        Returns:
            str: Chemin du fichier créé
        """
        if not filename:
            filename = self._get_filename()
        
        try:
            mode = 'a' if append else 'w'
            with open(filename, mode, newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                # Écrire les en-têtes seulement si c'est un nouveau fichier
                if not append:
                    writer.writerow(DEFAULT_HEADERS)
                
                # Écrire les données
                for ticket in tickets:
                    writer.writerow(self._ticket_to_row(ticket))
            
            action = "ajoutés à" if append else "exportés vers"
            logger.info(f"✅ {len(tickets)} tickets {action} {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export CSV: {e}")
            raise
    
    def export_to_excel(self, tickets: List[Dict], filename: Optional[str] = None) -> str:
        """
        Exporte les tickets vers un fichier Excel.
        
        Args:
            tickets: Liste de tickets normalisés
            filename: Nom de fichier personnalisé (optionnel)
            
        Returns:
            str: Chemin du fichier créé
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
            )
        
        if not filename:
            filename = self._get_filename()
        
        try:
            # Créer un nouveau workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tickets"
            
            # Écrire les en-têtes
            ws.append(DEFAULT_HEADERS)
            
            # Style les en-têtes (gras, fond gris)
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Écrire les données
            for ticket in tickets:
                ws.append(self._ticket_to_row(ticket))
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Max 50 caractères
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder le fichier
            wb.save(filename)
            
            logger.info(f"✅ {len(tickets)} tickets exportés vers {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export Excel: {e}")
            raise
    
    def export(self, tickets: List[Dict], filename: Optional[str] = None) -> str:
        """
        Exporte les tickets selon le format configuré.
        
        Args:
            tickets: Liste de tickets normalisés
            filename: Nom de fichier personnalisé (optionnel)
            
        Returns:
            str: Chemin du fichier créé
        """
        if not tickets:
            logger.warning("Aucun ticket à exporter")
            return None
        
        if self.file_format == "csv":
            return self.export_to_csv(tickets, filename)
        elif self.file_format == "xlsx":
            return self.export_to_excel(tickets, filename)
        else:
            raise ValueError(f"Format non supporté: {self.file_format}")
    
    def export_incremental(self, tickets: List[Dict], base_filename: str = "tickets") -> str:
        """
        Exporte les tickets avec un nom de fichier basé sur la date.
        Utile pour les exports incrémentaux quotidiens.
        
        Args:
            tickets: Liste de tickets normalisés
            base_filename: Nom de base du fichier
            
        Returns:
            str: Chemin du fichier créé
        """
        date_str = datetime.now().strftime("%Y%m%d")
        extension = "csv" if self.file_format == "csv" else "xlsx"
        filename = self.output_dir / f"{base_filename}_{date_str}.{extension}"
        
        # Si le fichier existe déjà, on ajoute les nouvelles données
        if filename.exists() and self.file_format == "csv":
            # Pour CSV, on ajoute les nouvelles lignes (sans les en-têtes)
            try:
                with open(filename, 'a', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    for ticket in tickets:
                        writer.writerow(self._ticket_to_row(ticket))
                logger.info(f"✅ {len(tickets)} tickets ajoutés à {filename}")
                return str(filename)
            except Exception as e:
                logger.error(f"Erreur lors de l'ajout au fichier CSV: {e}")
                raise
        else:
            # Nouveau fichier ou format Excel (Excel ne supporte pas l'ajout facilement)
            return self.export(tickets, str(filename))

    def merge_incremental_into_all(self, tickets: List[Dict]) -> str:
        """
        Fusionne les tickets (mis à jour ou nouveaux) dans tickets_all.csv.
        Par ticket_id : met à jour la ligne existante ou ajoute une nouvelle.
        Préserve l'ordre par ticket_id (numérique).
        """
        path = self.get_all_tickets_path()
        existing: Dict[str, list] = {}
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)  # skip header
                for row in reader:
                    if row and str(row[0]).strip():
                        existing[str(row[0]).strip()] = row
        for t in tickets:
            tid = str(t.get("ticket_id", "")).strip()
            existing[tid] = self._ticket_to_row(t)
        # Ordre : ticket_id numérique croissant, puis non-numériques
        def row_sort_key(row: list):
            if not row:
                return (1, "")
            try:
                return (0, int(row[0]))
            except (ValueError, TypeError):
                return (1, str(row[0]))

        sorted_rows = sorted(existing.values(), key=row_sort_key)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(DEFAULT_HEADERS)
            writer.writerows(sorted_rows)
        logger.info(f"tickets_all.csv mis à jour : {len(existing)} lignes ({len(tickets)} tickets fusionnés)")
        return str(path)

